import requests
import openpyxl
import telebot
from telebot import types

bot = telebot.TeleBot('5123025838:AAFEW5qcrLje8AyNMOQcaxEgAS4tRIpJ-Xo')
url = 'https://mtuci.ru/upload/iblock/18d/3-kurs-_SiSS_11.03.02-Infokommunikatsionnye-tekhnologii-i-sistemy-svyazi-VI.xlsx'  # актуальная ссылка на расписание (файл формата .xlsx)
silka = requests.get(url)
with open('table.xlsx', 'wb') as file:
    file.write(silka.content)
table = openpyxl.open('table.xlsx', read_only=True)

# 												wb = load_workbook('table.xlsx')
# 												print(wb.get_sheet_names())

# Понедельник
# Print out values in column 2
day = []
timer = ['9:30-11:05', '11:20-12:55', '13:10-14:45', '15:25-17:00', '17:15-18:50']
# sheet = table.worksheets[]  #4
nedelya = 7


@bot.message_handler(commands=['start'])
def start(message):
    keyboard = types.ReplyKeyboardMarkup()
    keyboard.row('/Нечетная', '/Четная', '/БТС1901', '/БСС1901', '/БСС1902', '/БЗС1901', '/БЗС1902', '/БОС1901', '/БСУ1901',
                 '/Понедельник', '/Вторник', '/Среда', '/Четверг', '/Пятница', '/Суббота')
    bot.send_message(message.chat.id,
                     "Здравствуйте,я бот Ya_Ystal (timetable_Bot) !\nИспользуйте мои команды чтобы узнать расписание "
                     "своей группы",
                     reply_markup=keyboard)


@bot.message_handler(commands=['Нечетная'])
def NechetWeak(message):
    global week
    global column_lecturer
    week = 7
    column_lecturer = 6
    # bot.send_message(message.chat.id, 'Это работает' + weak)


@bot.message_handler(commands=['Четная'])
def ChetWeak(message):
    global week
    global column_lecturer
    week = 8
    column_lecturer = 9
    # bot.send_message(message.chat.id, weak)


@bot.message_handler(commands=['БТС1901'])
def BTS1901(message):
    global sheet
    sheet = []
    sheet = table.worksheets[0]


@bot.message_handler(commands=['БСС1901'])
def BSS1901(message):
    global sheet
    sheet = []
    sheet = table.worksheets[1]


@bot.message_handler(commands=['БСС1902'])
def BSS1902(message):
    global sheet
    sheet = []
    sheet = table.worksheets[2]


@bot.message_handler(commands=['БЗС1901'])
def BSS1901(message):
    global sheet
    sheet = []
    sheet = table.worksheets[3]

@bot.message_handler(commands=['БЗС1902'])
def BZS1902(message):
    global sheet
    sheet = []
    sheet = table.worksheets[4]


@bot.message_handler(commands=['БОС1901'])
def BOS1901(message):
    global sheet
    sheet = []
    sheet = table.worksheets[5]


@bot.message_handler(commands=['БСУ1901'])  # /Понедельник
def BSU1901(message):
    global sheet
    sheet = []
    sheet = table.worksheets[6]


@bot.message_handler(commands=['Понедельник'])  # /Понедельник
def mondayfun(message):
    day = []
    surname = []
    for i in range(14, 19):
        day.append(sheet.cell(row=i, column=week).value)
    for sname in range(14, 19):
        surname.append(sheet.cell(row=sname, column=column_lecturer).value)
    for j in range(0, 5):
        bot.send_message(message.chat.id,
                         'Время: ' + timer[j] + ' \n' + str(day[j]) + '\nПреподаватель: ' + str(surname[j]))


@bot.message_handler(commands=['Вторник'])  # /Вторник
def vtornikfun(message):
    day = []
    surname = []
    for i in range(20, 25):
        day.append(sheet.cell(row=i, column=week).value)
    for sname in range(20, 25):
        surname.append(sheet.cell(row=sname, column=column_lecturer).value)
    for j in range(0, 5):
        bot.send_message(message.chat.id,
                         'Время: ' + timer[j] + ' \n' + str(day[j]) + '\nПреподаватель: ' + str(surname[j]))


@bot.message_handler(commands=['Среда'])  # /Среда
def sredafun(message):
    day = []
    surname = []
    for i in range(26, 31):
        day.append(sheet.cell(row=i, column=week).value)
    for sname in range(26, 31):
        surname.append(sheet.cell(row=sname, column=column_lecturer).value)
    for j in range(0, 5):
        bot.send_message(message.chat.id,
                         'Время: ' + timer[j] + ' \n' + str(day[j]) + '\nПреподаватель: ' + str(surname[j]))


@bot.message_handler(commands=['Четверг'])  # /Четверг
def chetvergfun(message):
    day = []
    surname = []
    for i in range(32, 37):
        day.append(sheet.cell(row=i, column=week).value)
    for sname in range((32, 37)):
        surname.append(sheet.cell(row=sname, column=column_lecturer).value)
    for j in range(0, 5):
        bot.send_message(message.chat.id,
                         'Время: ' + timer[j] + ' \n' + str(day[j]) + '\nПреподаватель: ' + str(surname[j]))


@bot.message_handler(commands=['Пятница'])  # /Пятница
def pyatnicafun(message):
    day = []
    surname = []
    for i in range(38, 43):
        day.append(sheet.cell(row=i, column=week).value)
    for sname in range(38, 43):
        surname.append(sheet.cell(row=sname, column=column_lecturer).value)
    for j in range(0, 5):
        bot.send_message(message.chat.id,
                         'Время: ' + timer[j] + ' \n' + str(day[j]) + '\nПреподаватель: ' + str(surname[j]))


@bot.message_handler(commands=['Суббота'])  # /Суббота
def subfun(message):
    day = []
    surname = []
    for i in range(44, 49):
        day.append(sheet.cell(row=i, column=week).value)
    for sname in range(44, 49):
        surname.append(sheet.cell(row=sname, column=column_lecturer).value)
    for j in range(0, 5):
        bot.send_message(message.chat.id,
                         'Время: ' + timer[j] + ' \n' + str(day[j]) + '\nПреподаватель: ' + str(surname[j]))


# print(monday)
bot.polling()

# bot = telebot.TeleBot('5123025838:AAFEW5qcrLje8AyNMOQcaxEgAS4tRIpJ-Xo')
# url = 'https://mtuci.ru/upload/iblock/18d/3-kurs-_SiSS_11.03.02-Infokommunikatsionnye-tekhnologii-i-sistemy-svyazi-VI.xlsx '
# актуальная ссылка на расписание (файл формата .xlsx)
# silka = requests.get(url)
# with open('table.xlsx', 'wb') as file:
#    file.write(silka.content)
